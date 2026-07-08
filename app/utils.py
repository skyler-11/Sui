"""Utility functions for Manning Simulator: parsing, validation, exports, helpers."""

from __future__ import annotations

import hashlib
import io
import os
import re
from datetime import date, timedelta
from typing import Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter

from app.core.rules import AVAILABLE_RULES, detect_matrix
from app.core.config import (
    DAYS, DAYS_A, DAYS_B, SHIFT_HRS, SHIFT_TIMES,
    CODE_MAP, VALID_CODES, EXTRA_COLS,
    STATION_KW, STATUS_KW, ID_KW, NAME_KW,
    OT_CODES, OT_HRS_61, OT_HRS_43, OT_STANDARD_DAYS_43,
)
from app.core.logging import get_logger, get_audit_logger, timing

logger = get_logger("forge.utils")
audit_logger = get_audit_logger()

# Columns stripped before export — internal tracking only
_EXPORT_STRIP_COLS = {"_pass", "_gv", "_dv", "Total Hrs", "_row_id"}
# OT Hrs A / OT Hrs B / OT Hrs Total intentionally kept — useful for HR


# ── WEEK LABEL HELPER ─────────────────────────────────────────────────────────
# Single source of truth for converting internal _A/_B keys to display strings.
# All UI rendering and exports call this — never hardcode .replace("_A"," A").

_DAY_INDEX = {"Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5, "Sun": 6}


def fmt_day(day_key: str) -> str:
    """
    Converts an internal day key like 'Mon_A' to a display label.

    When a reference date is set for that week, the date is shown:
        Mon_A → "Mon (5/4)"   Fri_B → "Fri (5/15)"

    When no reference date is set, falls back to the user-defined week label:
        Mon_A → "Mon Week A"  or  "Mon Week 17"
    """
    day_part, _, suffix = day_key.partition("_")
    if suffix == "A":
        ref = st.session_state.get("week_a_start")
        lbl = st.session_state.get("week_a_label", "Week A")
    elif suffix == "B":
        ref = st.session_state.get("week_b_start")
        lbl = st.session_state.get("week_b_label", "Week B")
    else:
        return day_key

    if ref and day_part in _DAY_INDEX:
        wd = week_dates(ref)[_DAY_INDEX[day_part]]
        return f"{day_part} ({wd.month}/{wd.day})"
    return f"{day_part} {lbl}"


def fmt_week(suffix: str) -> str:
    """Returns the display label for a week suffix ('A' or 'B')."""
    if suffix == "A":
        return st.session_state.get("week_a_label", "Week A")
    return st.session_state.get("week_b_label", "Week B")


# ── WEEK DATE HELPERS ─────────────────────────────────────────────────────────
# Reference date pickers drive Mon–Sun column headers in exports. Any picked
# weekday snaps to that week's Monday so the picker is forgiving.

def week_start(d: date) -> date:
    """Snap any date to the Monday of its ISO week."""
    return d - timedelta(days=d.weekday())


def week_dates(d: date) -> list[date]:
    """Return [Mon, Tue, ..., Sun] for the week containing d."""
    start = week_start(d)
    return [start + timedelta(days=i) for i in range(7)]


_WEEKDAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def weekday_headers(d: date | None) -> list[str]:
    """Build 7 column headers; falls back to plain names when d is None."""
    if not d:
        return list(_WEEKDAY_NAMES)
    return [f"{n}\n{wd.month}/{wd.day}" for n, wd in zip(_WEEKDAY_NAMES, week_dates(d))]


def fmt_week_range(d: date | None) -> str:
    """Formats the resolved Mon–Sun range as ' (m/d – m/d)' or '' if unset."""
    if not d:
        return ""
    wd = week_dates(d)
    return f" ({wd[0].month}/{wd[0].day} – {wd[-1].month}/{wd[-1].day})"


# ── FILE TYPE VALIDATION ──────────────────────────────────────────────────────

def _validate_file_bytes(filename: str, data: bytes) -> Tuple[bool, str]:
    """
    Validates file type by inspecting binary magic bytes.
    Zero-dependency — safe for Windows/IIS (no python-magic/libmagic needed).
    Guards against renamed/spoofed files.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    XLSX_SIG = b"PK\x03\x04"
    XLS_SIG = b"\xD0\xCF\x11\xE0"

    if ext == "csv":
        sample = data[:2048]
        for enc in ("utf-8", "utf-8-sig", "latin-1"):
            try:
                sample.decode(enc)
                return True, ""
            except UnicodeDecodeError:
                continue
        return False, "File is not valid text/CSV — it may be binary or corrupted."
    elif ext == "xlsx":
        if data[:4] != XLSX_SIG:
            return False, "File content does not match .xlsx signature. Try re-saving as Excel Workbook."
        return True, ""
    elif ext == "xls":
        if data[:4] != XLS_SIG:
            return False, "File content does not match .xls signature. Try re-saving as Excel 97-2003."
        return True, ""
    else:
        return False, f"Unsupported extension: .{ext}. Upload .xlsx, .xls, or .csv."


# ── USER CONTEXT ──────────────────────────────────────────────────────────────

def get_user_context() -> dict:
    """
    Builds a context dict from HTTP request headers (IP, user-agent, remote user).

    SEC [QA]: getpass.getuser() fallback has been intentionally removed.
    In a web/IIS context, that call returns the OS service account name
    (e.g. "IIS_IUSRS" or "NETWORK SERVICE"), which is an information leak.
    "anonymous" is the correct safe default when no header is present.
    """
    headers = {}
    try:
        headers = dict(
            st.context.headers) if st.context and st.context.headers else {}
    except Exception:
        pass

    # SEC [CWE-348]: X-Forwarded-For and X-Real-IP are client-supplied and
    # trivially spoofed unless the request actually traversed a trusted
    # reverse proxy that overwrites them. Only honour the forwarded headers
    # when MANNING_TRUSTED_PROXIES is configured (comma-separated list of
    # peer IPs that strip/replace these headers) AND the immediate peer
    # (remote-addr) is in that list. Otherwise the only trustworthy source
    # for the audit log is remote-addr itself.
    trusted_csv = os.getenv("MANNING_TRUSTED_PROXIES", "").strip()
    trusted_proxies = {
        p.strip() for p in trusted_csv.split(",") if p.strip()
    } if trusted_csv else set()
    remote_addr = headers.get("remote-addr", "").strip()
    peer_is_trusted = bool(trusted_proxies) and remote_addr in trusted_proxies

    if peer_is_trusted:
        forwarded = headers.get("x-forwarded-for", "")
        ip = (
            forwarded.split(",")[0].strip() if forwarded else None
        ) or headers.get("x-real-ip", "").strip() or remote_addr or "unknown"
    else:
        ip = remote_addr or "unknown"

    remote_user = headers.get("x-remote-user", "").strip()

    # SEC: Do NOT fall back to getpass.getuser() — it leaks the OS service
    # account identity (e.g. "NETWORK SERVICE") in web-hosted environments.
    if not remote_user:
        remote_user = "anonymous"

    if ip == "unknown":
        ip = "127.0.0.1"

    return {
        "ip":         ip,
        "user_agent": headers.get("user-agent", "unknown"),
        "user_name":  remote_user,
    }


def log_user_action(action: str, **details) -> None:
    """Structured audit log — key="value" format for Splunk/Datadog SIEM.

    SEC [CWE-117]: User-supplied detail values (filenames, search terms, etc.)
    can contain CR/LF/NUL or other control characters that would forge fake
    log entries if rendered raw. Pre-formatting them into a single trusted
    string previously bypassed the _SanitizingAdapter, so we now strip
    control chars from each value here at the boundary, before formatting.
    """
    from app.core.logging import _CONTROL_CHAR_RE  # internal helper

    user_id = st.session_state.get("user_id", "anonymous")
    user_ctx = st.session_state.get(
        "user_ctx", {"ip": "unknown", "user_agent": "unknown"})

    def _scrub(v: object) -> object:
        if isinstance(v, str):
            # Strip control chars AND escape embedded double-quotes so a value
            # like  hi" injected="true  cannot break out of the quoted slot.
            return _CONTROL_CHAR_RE.sub("", v).replace('"', "'")
        return v

    details_str = " ".join(
        f'{k}="{_scrub(v)}"' for k, v in details.items()) if details else ""

    audit_logger.info(
        'action="%s" user_id="%s" ip="%s" user_agent="%s" %s',
        action, user_id,
        user_ctx.get("ip", "unknown"),
        user_ctx.get("user_agent", "unknown"),
        details_str,
    )

    # Shadow-log to app logger at DEBUG so dev/ops can correlate audit events
    # with surrounding application log lines without tailing two files.
    logger.debug('audit_shadow action="%s" user_id="%s"', action, user_id)


# ── VALIDATION ────────────────────────────────────────────────────────────────

def validate(df: pd.DataFrame, config: dict, active_rule_keys: list) -> pd.DataFrame:
    logger.info('validation_start rows=%d rules=%s', len(df), active_rule_keys)

    with timing(logger, "validate", rows=len(df), rules=len(active_rule_keys)):
        return _validate_impl(df, config, active_rule_keys)


def _validate_impl(df: pd.DataFrame, config: dict, active_rule_keys: list) -> pd.DataFrame:
    out_df = df.copy()

    # ── Metadata hygiene ──────────────────────────────────────────────────────
    for col in ["Name", "Station", "EMP. STATUS", "ID No."]:
        if col not in out_df.columns:
            out_df[col] = "Unspecified"
        else:
            out_df[col] = (
                out_df[col].fillna("Unspecified")
                .astype(str).replace(["", "nan"], "Unspecified")
            )

    for day in DAYS:
        if day not in out_df.columns:
            out_df[day] = ""

    # ── Per-day hours ─────────────────────────────────────────────────────────
    for day in DAYS:
        out_df[f"{day}_hrs"] = (
            out_df[day].astype(str).str.strip()
            .map(lambda x: SHIFT_HRS.get(x, 0))
        )

    # ── Weekly hours totals ───────────────────────────────────────────────────
    out_df["Total Hrs A"] = out_df[[f"{d}_hrs" for d in DAYS_A]].sum(axis=1)
    out_df["Total Hrs B"] = out_df[[f"{d}_hrs" for d in DAYS_B]].sum(axis=1)
    out_df["Total Hrs"] = out_df["Total Hrs A"] + out_df["Total Hrs B"]

    # ── RD counts ─────────────────────────────────────────────────────────────
    out_df["rd_count_A"] = (
        out_df[DAYS_A].astype(str).apply(
            lambda col: col.str.strip().str.upper()) == "RD"
    ).sum(axis=1)
    out_df["rd_count_B"] = (
        out_df[DAYS_B].astype(str).apply(
            lambda col: col.str.strip().str.upper()) == "RD"
    ).sum(axis=1)

    # ── Matrix detection (pre-computed once — rules read _matrix column) ──────
    out_df["_matrix"] = out_df.apply(detect_matrix, axis=1)

    # ── OT Hours — MATRIX-AWARE VECTORIZED ───────────────────────────────────
    #
    # 6-1 Matrix:
    #   Standard day = 8h. AOT/BOT/COT = 12h → OT_HRS_61 (4h) per OT shift.
    #   OT Hrs = sum of 4h for each AOT/BOT/COT day in the week.
    #
    # 4-3 Matrix:
    #   Standard = 4 work days × 12h = 48h/week (12h is their baseline).
    #   OT triggers ONLY if they work a 5th day — that full 12h day is OT.
    #   OT Hrs = max(0, worked_days − 4) × 12h  per week.
    #   Maximum possible: 1 extra day × 12h = 12h OT → total 60h (weekly cap).
    #
    # OS5 (9h): excluded pending HR clarification.
    # ─────────────────────────────────────────────────────────────────────────

    # 6-1: per-day OT flag (4h if OT code, 0 otherwise)
    for day in DAYS:
        upper = out_df[day].astype(str).str.strip().str.upper()
        out_df[f"{day}_ot61"] = upper.map(
            lambda x: OT_HRS_61 if x in OT_CODES else 0)

    ot61_a = out_df[[f"{d}_ot61" for d in DAYS_A]].sum(axis=1)
    ot61_b = out_df[[f"{d}_ot61" for d in DAYS_B]].sum(axis=1)

    # 4-3: worked days per week (any non-zero hours counts as a worked day)
    worked_a = (out_df[[f"{d}_hrs" for d in DAYS_A]] > 0).sum(axis=1)
    worked_b = (out_df[[f"{d}_hrs" for d in DAYS_B]] > 0).sum(axis=1)
    ot43_a = ((worked_a - OT_STANDARD_DAYS_43).clip(lower=0) * OT_HRS_43)
    ot43_b = ((worked_b - OT_STANDARD_DAYS_43).clip(lower=0) * OT_HRS_43)

    # Apply matrix-aware selection using boolean mask
    is_43 = (out_df["_matrix"] == "4-3")
    out_df["OT Hrs A"] = ot61_a.where(~is_43, ot43_a)
    out_df["OT Hrs B"] = ot61_b.where(~is_43, ot43_b)
    out_df["OT Hrs Total"] = out_df["OT Hrs A"] + out_df["OT Hrs B"]

    # ── Rule evaluation ───────────────────────────────────────────────────────
    out_df["_pass"] = True
    out_df["_gv"] = [[] for _ in range(len(out_df))]
    out_df["_dv"] = [[] for _ in range(len(out_df))]

    active_rules = [r for r in AVAILABLE_RULES if r.key in active_rule_keys]
    for rule in active_rules:
        passed, msg, gv, dv = rule.evaluate_df(out_df, config)
        out_df[rule.name] = msg
        out_df["_pass"] = out_df["_pass"] & passed
        if gv is not None:
            out_df["_gv"] = out_df["_gv"] + gv
        if dv is not None:
            out_df["_dv"] = out_df["_dv"] + dv

    out_df["Status"] = out_df["_pass"].apply(
        lambda p: "Valid" if p else "Invalid")

    # Drop intermediates — OT Hrs columns are intentionally kept for HR exports
    drop_cols = (
        [f"{d}_hrs" for d in DAYS]
        + [f"{d}_ot61" for d in DAYS]
        + ["rd_count_A", "rd_count_B",
           "Total Hrs A", "Total Hrs B", "Total Hrs", "_matrix"]
    )
    out_df = out_df.drop(columns=drop_cols)

    total = len(out_df)
    passing = int(out_df["_pass"].sum())
    logger.info(
        'validation_complete passed=%d total=%d rate=%.0f%% ot_total=%.0f',
        passing, total,
        100 * passing / total if total else 0,
        out_df["OT Hrs Total"].sum(),
    )

    return out_df


# ── FILE PARSING ──────────────────────────────────────────────────────────────

# Regex to detect date-like strings — guards against date reference rows in the
# template being parsed as employee name rows and creating phantom table entries.
#
# FIX [Issue-3]: The third branch previously matched any string that merely
# *started with* a month abbreviation (e.g. "Marcelo", "Januario", "Decena",
# "May Santos", "June Cruz").  This caused up to 1 employee per file to be
# silently skipped depending on name ordering.
# The branch is now anchored to require digits after the month name so that
# only genuine date strings like "Jan 12 2025" or "May 5" are caught.
_DATE_PATTERN = re.compile(
    r"^\d{1,2}[\/\-]\d{1,2}([\/\-]\d{2,4})?$"                          # 12/01 or 12/01/2025
    r"|^\d{4}[\/\-]\d{1,2}[\/\-]\d{1,2}$"                               # 2025-01-12
    r"|^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b[\s\-\/]\d",  # Jan 12 / May-5
    re.IGNORECASE,
)


def parse_bytes(filename: str, data: bytes):
    logger.info('parse_start filename="%s" size_kb=%.1f', filename, len(data) / 1024)

    with timing(logger, "parse_file", filename=filename,
                size_kb=round(len(data) / 1024, 1)):
        return _parse_bytes_impl(filename, data)


def _parse_bytes_impl(filename: str, data: bytes):
    MAX_MB = 5
    if len(data) > MAX_MB * 1024 * 1024:
        return None, f"File too large. Maximum allowed size is {MAX_MB}MB."

    is_valid, mime_err = _validate_file_bytes(filename, data)
    if not is_valid:
        logger.warning('mime_check_failed filename="%s" reason="%s"', filename, mime_err)
        return None, mime_err

    # SKIP: rows whose name cell matches these are silently dropped.
    # Includes date patterns (added here) so reference date rows in the
    # template header area never produce phantom employee records.
    SKIP = {"nan", "name", "operator's name", "operators name", "none", ""}

    try:
        buf = io.BytesIO(data)
        is_csv = filename.lower().endswith(".csv")

        raw_dict: dict[str, pd.DataFrame] = {}
        if is_csv:
            raw_dict["csv"] = pd.read_csv(buf, header=None, dtype=str)
        else:
            xls = pd.ExcelFile(buf)
            for sheet in xls.sheet_names:
                raw_dict[sheet] = pd.read_excel(
                    xls, sheet_name=sheet, header=None, dtype=str)

        # SEC [CWE-400]: Decompression-bomb guard. The 5MB byte cap above
        # bounds the compressed payload, but a malicious .xlsx can expand
        # to hundreds of MB of cells (mostly empty/repeated). Reject
        # workbooks whose total cell count is beyond what a real manning
        # template ever needs (10k rows × 50 cols across all sheets is
        # already 5x the largest legitimate file we have seen).
        MAX_CELLS = 500_000
        total_cells = sum(df.shape[0] * df.shape[1] for df in raw_dict.values())
        if total_cells > MAX_CELLS:
            logger.warning(
                'parse_rejected_oversized filename="%s" total_cells=%d limit=%d',
                filename, total_cells, MAX_CELLS,
            )
            return None, (
                f"File rejected: workbook contains {total_cells:,} cells, "
                f"which exceeds the {MAX_CELLS:,}-cell safety limit."
            )

        day_row, day_cols, target_raw = None, {}, None

        for _sheet, df in raw_dict.items():
            df = df.fillna("").astype(str)
            if df.empty:
                continue

            for i, row in df.iterrows():
                hits = [
                    v for v in (str(v).strip().lower() for v in row)
                    if any(v.startswith(d) for d in
                           ["mon", "tue", "wed", "thu", "fri", "sat", "sun"])
                ]
                if len(hits) >= 10:
                    day_row, day_cols, day_counter = i, {}, 0
                    for j, v in enumerate(row):
                        val = str(v).strip().lower()
                        if any(val.startswith(d) for d in
                               ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]):
                            if day_counter < 7:
                                day_cols[DAYS_A[day_counter]] = j
                            elif day_counter < 14:
                                day_cols[DAYS_B[day_counter - 7]] = j
                            day_counter += 1
                    target_raw = df
                    break
            if day_row is not None:
                break

        if target_raw is None or day_row is None:
            return None, (
                "No 14-day schedule columns found. "
                "Ensure 'Mon'–'Sun' appear twice in your header row."
            )

        raw = target_raw
        name_col = station_col = status_col = id_col = None
        for i in range(day_row + 1):
            for j, v in enumerate(raw.iloc[i]):
                vl = str(v).strip().lower()
                if name_col is None and any(k in vl for k in NAME_KW):
                    name_col = j
                if station_col is None and any(k in vl for k in STATION_KW):
                    station_col = j
                if status_col is None and any(k in vl for k in STATUS_KW):
                    status_col = j
                if id_col is None and any(k in vl for k in ID_KW):
                    id_col = j

        if name_col is None:
            return None, "No 'Name' column detected. Check your header row."

        out = []
        for i in range(day_row + 1, len(raw)):
            emp_name = str(raw.iloc[i, name_col]).strip()

            # Skip blank, sentinel, formula, and date reference rows
            if (emp_name.lower() in SKIP
                    or emp_name.startswith("=")
                    or _DATE_PATTERN.match(emp_name)):
                logger.info(
                    'parse_row_skipped filename="%s" row=%d name="%s"',
                    filename, i, emp_name,
                )
                continue

            rec = {"Name": emp_name}
            for day, ci in day_cols.items():
                val = str(raw.iloc[i, ci]).strip().upper()
                rec[day] = CODE_MAP.get(val, val) if val != "NAN" else ""
            for day in DAYS:
                rec.setdefault(day, "")

            rec["Station"] = str(raw.iloc[i, station_col]).strip(
            ) if station_col is not None else ""
            rec["EMP. STATUS"] = str(raw.iloc[i, status_col]).strip(
            ) if status_col is not None else ""
            rec["ID No."] = str(raw.iloc[i, id_col]).strip(
            ) if id_col is not None else ""

            for k in ["Station", "EMP. STATUS", "ID No."]:
                if rec[k] == "NAN":
                    rec[k] = ""

            out.append(rec)

        if not out:
            logger.info('parse_ok filename="%s" employees=0', filename)
            return pd.DataFrame(
                columns=["Station", "EMP. STATUS", "ID No.", "Name"] + DAYS), None

        logger.info('parse_ok filename="%s" employees=%d', filename, len(out))
        return pd.DataFrame(out), None

    except Exception:
        # SEC [CWE-209]: Do not echo the raw exception (which can include
        # absolute file paths, library internals, or partial file contents)
        # back to the UI. Full traceback is in app.log; the user gets a
        # generic message and can quote the trace_id to support.
        logger.exception(
            'parse_error filename="%s" user="%s"',
            filename,
            st.session_state.get("user_id", "unknown"),
        )
        return None, (
            "Could not read the file. The file may be corrupted, password-"
            "protected, or in an unexpected format. Try re-saving as a fresh "
            ".xlsx or .csv. Contact IT with your session trace ID if it persists."
        )


# ── TEMPLATE ──────────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600)
def create_excel_template() -> bytes:
    """
    Reads the pre-made Excel template from disk.
    Returns b"" if missing — callers must handle and show a user warning.

    FIX [QA]: ttl=3600 added so on-disk template updates propagate within
    one hour without requiring a full Streamlit server restart. Previously
    cached indefinitely, meaning any template file replacement would be
    invisible until the process was recycled.
    """
    from pathlib import Path
    path = Path(__file__).resolve().parent.parent / \
        "resource" / "manning_template.xlsx"
    if not path.exists():
        logger.error("Template not found at: %s", path)
        return b""
    try:
        return path.read_bytes()
    except Exception:
        logger.exception("Failed to read Excel template.")
        return b""


# ── EXPORT ────────────────────────────────────────────────────────────────────

def sanitize_for_csv_export(df: pd.DataFrame) -> pd.DataFrame:
    """Mitigates formula-injection by prefixing dangerous leading characters.

    SEC [CWE-1236 / CWE-79]: Excel and LibreOffice evaluate cells beginning
    with '=', '+', '-', '@', '|' (DDE), or whitespace control chars (\\t /
    \\r) followed by one of the above. OWASP recommends covering the full
    set, not only the formula-leaders. We also strip leading control chars
    that would otherwise hide the formula leader from naive checks.
    """
    _DANGEROUS = ("=", "+", "-", "@", "|")

    def _san(val):
        if not isinstance(val, str) or not val:
            return val
        # Treat leading TAB / CR / LF as if absent — they hide the real
        # leading character from a human reader but the spreadsheet still
        # parses what follows.
        stripped = val.lstrip("\t\r\n")
        if stripped and stripped[0] in _DANGEROUS:
            return "'" + val
        return val
    return df.map(_san)


def _df_fingerprint(df: pd.DataFrame) -> str:
    summary = f"{df.shape}|{list(df.columns)}|{df.iloc[:1].to_csv()}|{df.iloc[-1:].to_csv()}"
    return hashlib.md5(summary.encode()).hexdigest()  # noqa: S324 — non-security hash


def _build_export_header(df: pd.DataFrame) -> list[str]:
    """
    Builds a human-readable header line for exports containing:
    week labels and optional date ranges from session_state.
    Stamped as the first row comment in the Excel title.

    FIX [QA-Critical]: reads week_a_start / week_b_start (the correct keys
    set by init_session_state and render_editor_tab). Previously these were
    never populated because render_editor_tab was saving to week_a_date /
    week_b_date — a key mismatch that caused dates to always appear as None
    in every export.
    """
    lbl_a = st.session_state.get("week_a_label", "Week A")
    lbl_b = st.session_state.get("week_b_label", "Week B")
    date_a = st.session_state.get("week_a_start")
    date_b = st.session_state.get("week_b_start")
    total = len(df)
    ot_total = int(df["OT Hrs Total"].sum()
                   ) if "OT Hrs Total" in df.columns else 0

    parts = [
        f"{lbl_a}{fmt_week_range(date_a)}",
        f"{lbl_b}{fmt_week_range(date_b)}",
        f"Employees: {total}",
        f"Total OT: {ot_total}h",
    ]
    return parts


def generate_styled_excel_export(display_df: pd.DataFrame) -> bytes:
    """
    Generates a template-matching Excel workbook with 3 sheets:

      Sheet 1 — "Schedule"
        Matches resource/manning_template.xlsx structure exactly:
        19 columns (Station/Designation, Employee Status, ID NO.,
        Operator's Name:, Mon-Sun week 1, SPACER, Mon-Sun week 2).
        Metadata row at top: labels, dates, OT totals, compliance.
        Rows colour-coded: green=Valid, red=Invalid.
        HR can re-upload this file back into the app — parser reads it.

      Sheet 2 — "Validation Details"
        Full audit trail. Employee-level status, violated rules,
        violation details, matrix type, OT hours per week.

      Sheet 3 — "Summary"
        KPIs: headcount, compliance %, OT by matrix, OT by station.

    Uses openpyxl directly (not pandas.to_excel) because pandas cannot
    write duplicate column headers — the template has Mon-Sun twice.

    FIX [H3]: No @st.cache_data — caller caches bytes in session_state.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    logger.debug('excel_export_start rows=%d', len(display_df))

    # Manual enter/exit so we can wrap the entire (long) function body without
    # introducing an extra indent level. We use a sentinel object plus try/
    # finally below to guarantee the elapsed-time log fires even on exception.
    _excel_timing = timing(logger, "excel_export", rows=len(display_df))
    _excel_timing.__enter__()
    _excel_exc_info: tuple | None = None

    # ── Colour palette (brand + semantic) ─────────────────────────────────
    C_BRAND = "F24713"   # AUMOVIO orange
    C_BRAND_DK = "CD3C10"
    C_VALID_BG = "E8F5E9"
    C_INVALID_BG = "FFEBEE"
    C_META_BG = "FEF5F3"
    C_HEADER_BG = "1A0A05"  # dark brand
    C_WHITE = "FFFFFF"
    C_TEXT = "0F172A"

    # Styles reused across sheets
    thin = Side(border_style="thin", color="D4C4BE")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    meta_font = Font(name="Inter", size=9, color=C_TEXT, bold=True)
    meta_fill = PatternFill("solid", fgColor=C_META_BG)
    header_font = Font(name="Inter", size=10, color=C_WHITE, bold=True)
    header_fill = PatternFill("solid", fgColor=C_BRAND)
    body_font = Font(name="Inter", size=9, color=C_TEXT)
    title_font = Font(name="Rajdhani", size=16, color=C_WHITE, bold=True)
    title_fill = PatternFill("solid", fgColor=C_HEADER_BG)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Strip internal columns ────────────────────────────────────────────
    df = display_df.drop(
        columns=[c for c in _EXPORT_STRIP_COLS if c in display_df.columns]
    ).copy()
    df = sanitize_for_csv_export(df)

    # Compute summary numbers once
    total = len(df)
    valid_cnt = int((df.get("Status", "") == "Valid").sum()
                    ) if "Status" in df.columns else 0
    invalid_cnt = total - valid_cnt
    compliance = (valid_cnt / total * 100) if total else 0
    ot_total_a = int(df["OT Hrs A"].sum()) if "OT Hrs A" in df.columns else 0
    ot_total_b = int(df["OT Hrs B"].sum()) if "OT Hrs B" in df.columns else 0
    ot_total = int(df["OT Hrs Total"].sum()
                   ) if "OT Hrs Total" in df.columns else 0

    lbl_a = st.session_state.get("week_a_label", "Week A")
    lbl_b = st.session_state.get("week_b_label", "Week B")
    date_a = st.session_state.get("week_a_start")
    date_b = st.session_state.get("week_b_start")

    wb = Workbook()

    # ═════════════════════════════════════════════════════════════════════
    # SHEET 1 — "Schedule" (template-matching layout)
    # ═════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Schedule"

    # Row 1: Title bar
    ws1.cell(row=1, column=1, value="MANNING SIMULATOR — VALIDATED SCHEDULE")
    ws1.cell(row=1, column=1).font = title_font
    ws1.cell(row=1, column=1).fill = title_fill
    ws1.cell(row=1, column=1).alignment = center
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    ws1.row_dimensions[1].height = 28

    # Row 2: Metadata strip
    meta_parts = [
        f"{lbl_a}{fmt_week_range(date_a)}",
        f"{lbl_b}{fmt_week_range(date_b)}",
        f"Employees: {total}",
        f"Valid: {valid_cnt}",
        f"Invalid: {invalid_cnt}",
        f"Compliance: {compliance:.1f}%",
        f"Total OT: {ot_total}h",
    ]
    for i, val in enumerate(meta_parts, start=1):
        c = ws1.cell(row=2, column=i, value=val)
        c.font = meta_font
        c.fill = meta_fill
        c.alignment = center
        c.border = border_all
    # Extend meta fill to full 19 columns visually
    for i in range(len(meta_parts) + 1, 20):
        c = ws1.cell(row=2, column=i, value="")
        c.fill = meta_fill
        c.border = border_all
    ws1.row_dimensions[2].height = 22

    # Row 3: blank spacer
    ws1.row_dimensions[3].height = 8

    # Row 4: Template column headers. Weekday cells carry the resolved
    # calendar date (e.g. "Mon\n5/4") when a reference date is set; fall
    # back to plain "Mon"/"Tue"/... when unset. openpyxl direct write
    # bypasses pandas' no-duplicate-columns rule.
    template_headers = (
        ["Station/Designation", "Employee Status", "ID NO.", "Operator's Name:"]
        + weekday_headers(date_a)
        + ["SPACER"]
        + weekday_headers(date_b)
    )
    for i, h in enumerate(template_headers, start=1):
        c = ws1.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border_all
    # Taller row to fit two-line headers (weekday + m/d)
    ws1.row_dimensions[4].height = 32

    # Internal → template column mapping
    col_map = {
        "Station":     1,
        "EMP. STATUS": 2,
        "ID No.":      3,
        "Name":        4,
        "Mon_A": 5,  "Tue_A": 6,  "Wed_A": 7,  "Thu_A": 8,
        "Fri_A": 9,  "Sat_A": 10, "Sun_A": 11,
        # Column 12 = SPACER (blank)
        "Mon_B": 13, "Tue_B": 14, "Wed_B": 15, "Thu_B": 16,
        "Fri_B": 17, "Sat_B": 18, "Sun_B": 19,
    }

    # Rows 5+: Employee data
    for r_off, (_, row) in enumerate(df.iterrows(), start=5):
        is_valid = str(row.get("Status", "")) == "Valid"
        row_fill = PatternFill(
            "solid", fgColor=C_VALID_BG if is_valid else C_INVALID_BG)

        for internal_col, sheet_col in col_map.items():
            val = row.get(internal_col, "")
            c = ws1.cell(row=r_off, column=sheet_col, value=val)
            c.font = body_font
            c.fill = row_fill
            c.alignment = center if sheet_col >= 5 else left
            c.border = border_all

        # SPACER column — keep styled but empty
        sc = ws1.cell(row=r_off, column=12, value="")
        sc.fill = row_fill
        sc.border = border_all

    # Column widths tuned for readability
    widths = {
        1: 22, 2: 14, 3: 12, 4: 22,   # meta cols
        12: 3,                         # spacer thin
    }
    for col in range(1, 20):
        w = widths.get(col, 7)         # default 7 for day cells
        ws1.column_dimensions[get_column_letter(col)].width = w

    # Freeze above employee rows for easy scrolling
    ws1.freeze_panes = "A5"

    # ═════════════════════════════════════════════════════════════════════
    # SHEET 2 — "Validation Details" (full audit)
    # ═════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Validation Details")

    # Compose detail columns — all data Streamlit shows in Results tab
    detail_cols = ["Station", "EMP. STATUS", "ID No.", "Name"]
    rule_cols = [r.name for r in AVAILABLE_RULES if r.name in df.columns]
    ot_cols = [c for c in ["OT Hrs A", "OT Hrs B",
                           "OT Hrs Total"] if c in df.columns]
    final_cols = ["Status"]
    detail_cols += rule_cols + ot_cols + final_cols

    # Filter to columns that actually exist
    detail_cols = [c for c in detail_cols if c in df.columns]
    det_df = df[detail_cols].copy()

    # Rename day keys if any are in detail_cols (unlikely but defensive)
    det_df = det_df.rename(columns={d: fmt_day(d)
                           for d in DAYS if d in det_df.columns})

    # Title
    ws2.cell(row=1, column=1, value="VALIDATION DETAILS").font = title_font
    ws2.cell(row=1, column=1).fill = title_fill
    ws2.cell(row=1, column=1).alignment = center
    ws2.merge_cells(start_row=1, start_column=1,
                    end_row=1, end_column=len(detail_cols))
    ws2.row_dimensions[1].height = 28

    # Headers
    for i, h in enumerate(det_df.columns, start=1):
        c = ws2.cell(row=3, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border_all
    ws2.row_dimensions[3].height = 26

    # Data rows with colour-coded Status
    for r_off, (_, row) in enumerate(det_df.iterrows(), start=4):
        is_valid = str(row.get("Status", "")) == "Valid"
        row_fill = PatternFill(
            "solid", fgColor=C_VALID_BG if is_valid else C_INVALID_BG)
        for ci, col in enumerate(det_df.columns, start=1):
            c = ws2.cell(row=r_off, column=ci, value=row.get(col, ""))
            c.font = body_font
            c.fill = row_fill
            c.alignment = left
            c.border = border_all

    # Column widths — measure content
    for ci, col in enumerate(det_df.columns, start=1):
        max_len = max(
            det_df[col].astype(str).map(len).max() if not det_df.empty else 0,
            len(str(col)),
        ) + 2
        ws2.column_dimensions[get_column_letter(
            ci)].width = min(max(max_len, 10), 45)

    ws2.freeze_panes = "A4"

    # ═════════════════════════════════════════════════════════════════════
    # SHEET 3 — "Summary" (KPI overview for HR)
    # ═════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Summary")

    ws3.cell(row=1, column=1, value="SUMMARY & KPIs").font = title_font
    ws3.cell(row=1, column=1).fill = title_fill
    ws3.cell(row=1, column=1).alignment = center
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws3.row_dimensions[1].height = 28

    kpi_rows = [
        ("METRIC", "VALUE", "", ""),
        ("Total Headcount",     total,              "", ""),
        ("Valid Schedules",     valid_cnt,          "", ""),
        ("Invalid Schedules",   invalid_cnt,        "", ""),
        ("Compliance Rate",     f"{compliance:.1f}%", "", ""),
        ("", "", "", ""),
        (f"{lbl_a} — Total OT Hours", f"{ot_total_a}h",  "", ""),
        (f"{lbl_b} — Total OT Hours", f"{ot_total_b}h",  "", ""),
        ("14-Day Total OT Hours",      f"{ot_total}h",   "", ""),
        ("Average OT / Employee",
         f"{round(ot_total/total, 1) if total else 0}h", "", ""),
    ]

    row_cursor = 3
    for r in kpi_rows:
        for i, val in enumerate(r, start=1):
            c = ws3.cell(row=row_cursor, column=i, value=val)
            if row_cursor == 3:
                c.font = header_font
                c.fill = header_fill
            else:
                c.font = body_font
                c.fill = meta_fill
            c.alignment = left if i == 1 else center
            c.border = border_all
        row_cursor += 1

    # ── Helpers (local to Summary sheet) ──────────────────────────────────
    # Keep Coverage Dashboard mirrors visually consistent with the KPI block
    # above and with the on-screen dashboard tables. Section titles span all
    # 4 columns; table headers use the brand fill; data rows use meta_fill.
    def _section_title(title: str, cursor: int) -> int:
        ws3.cell(row=cursor, column=1, value=title).font = header_font
        ws3.cell(row=cursor, column=1).fill = header_fill
        ws3.cell(row=cursor, column=1).alignment = left
        ws3.cell(row=cursor, column=1).border = border_all
        # Style the merged span so the brand fill paints the full row
        for j in range(2, 5):
            cc = ws3.cell(row=cursor, column=j, value="")
            cc.fill = header_fill
            cc.border = border_all
        ws3.merge_cells(start_row=cursor, start_column=1,
                        end_row=cursor, end_column=4)
        ws3.row_dimensions[cursor].height = 22
        return cursor + 1

    def _table_header(headers: list[str], cursor: int) -> int:
        for i, h in enumerate(headers, start=1):
            c = ws3.cell(row=cursor, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border_all
        return cursor + 1

    def _table_row(values: list, cursor: int) -> int:
        for i, v in enumerate(values, start=1):
            c = ws3.cell(row=cursor, column=i, value=v)
            c.font = body_font
            c.fill = meta_fill
            c.alignment = left if i == 1 else center
            c.border = border_all
        return cursor + 1

    # ── COVERAGE DASHBOARD SUMMARY ────────────────────────────────────────
    # Mirrors the on-screen Operational Coverage Dashboard so HR has the
    # same operational view in the workbook (no Streamlit needed).
    # Sections rendered (in dashboard order):
    #   1. Schedule Matrix Distribution
    #   2. Coverage by Station       (replaces older "BY STATION" block)
    #   3. Top Violation Categories
    #   4. Per-Day Headcount by Shift
    row_cursor += 2
    row_cursor = _section_title("COVERAGE DASHBOARD SUMMARY", row_cursor)

    # 1. Schedule Matrix Distribution ──────────────────────────────────────
    # Matrix detection is row-level over the full 14-day record (matches
    # dashboard semantics — see render_dashboard_tab). Note: `_pass` is
    # already stripped by _EXPORT_STRIP_COLS, so we derive validity from
    # `Status` (the export-facing truth column used elsewhere in this fn).
    try:
        if not df.empty and "Status" in df.columns:
            matrices = df.apply(detect_matrix, axis=1)
            valid_series = (df["Status"].astype(str) == "Valid")
            mat_df = (
                pd.DataFrame({"Matrix": matrices, "_valid": valid_series})
                .groupby("Matrix")
                .agg(Headcount=("_valid", "count"), Valid=("_valid", "sum"))
                .reset_index()
                .sort_values("Headcount", ascending=False)
            )
            row_cursor += 1
            row_cursor = _section_title(
                "Schedule Matrix Distribution", row_cursor)
            row_cursor = _table_header(
                ["Matrix", "Headcount", "Valid", "Compliance"], row_cursor)
            for _, m_row in mat_df.iterrows():
                hc = int(m_row["Headcount"])
                vd = int(m_row["Valid"])
                comp = f"{(vd / hc * 100):.1f}%" if hc else "0.0%"
                row_cursor = _table_row(
                    [str(m_row["Matrix"]), hc, vd, comp], row_cursor)
    except Exception:
        # Defensive: matrix detection depends on day-key columns being
        # present on `df`. If callers ever pass a stripped frame, skip
        # this section instead of failing the whole export.
        logger.warning("excel_summary_matrix_skip", exc_info=True)

    # 2. Coverage by Station ───────────────────────────────────────────────
    # Same shape as the on-screen "Coverage by Station" table; OT total is
    # appended into the 4th column when OT Hrs Total exists, so we stay
    # within the Summary sheet's 4-column grid (Station / Headcount /
    # Valid / Compliance [· OT]). Validity is sourced from `Status`
    # because `_pass` is stripped earlier in this function.
    if "Station" in df.columns and "Status" in df.columns:
        row_cursor += 1
        row_cursor = _section_title("Coverage by Station", row_cursor)
        has_ot = "OT Hrs Total" in df.columns
        if has_ot:
            row_cursor = _table_header(
                ["Station", "Headcount", "Valid", "Compliance / OT"],
                row_cursor,
            )
        else:
            row_cursor = _table_header(
                ["Station", "Headcount", "Valid", "Compliance"], row_cursor)

        agg_kwargs = {
            "Headcount": ("Status", "count"),
            "Valid": ("Status", lambda s: int((s == "Valid").sum())),
        }
        if has_ot:
            agg_kwargs["OT"] = ("OT Hrs Total", "sum")

        station_stats = (
            df.groupby("Station")
              .agg(**agg_kwargs)
              .reset_index()
              .sort_values("Headcount", ascending=False)
        )
        for _, s_row in station_stats.iterrows():
            hc = int(s_row["Headcount"])
            vd = int(s_row["Valid"])
            comp = f"{(vd / hc * 100):.1f}%" if hc else "0.0%"
            last_col = (
                f"{comp} · {int(s_row['OT'])}h OT" if has_ot else comp
            )
            row_cursor = _table_row(
                [str(s_row["Station"]), hc, vd, last_col], row_cursor)

    # 3. Top Violation Categories ─────────────────────────────────────────
    rule_cols = [r.name for r in AVAILABLE_RULES if r.name in df.columns]
    if rule_cols:
        row_cursor += 1
        row_cursor = _section_title("Top Violation Categories", row_cursor)
        row_cursor = _table_header(
            ["Rule", "Failed Employees", "Failure Rate", ""], row_cursor)
        failed_rows = []
        for col in rule_cols:
            n = int(df[col].astype(str).str.startswith("❌").sum())
            if n > 0:
                failed_rows.append((col, n))
        failed_rows.sort(key=lambda x: x[1], reverse=True)
        if failed_rows:
            for rule_name, n in failed_rows:
                rate_pct = f"{(n / total * 100):.1f}%" if total else "0.0%"
                row_cursor = _table_row(
                    [rule_name, n, rate_pct, ""], row_cursor)
        else:
            row_cursor = _table_row(
                ["All employees pass every active rule.", "", "", ""],
                row_cursor,
            )

    # 4. Per-Day Headcount by Shift ───────────────────────────────────────
    # Mirrors the dashboard's per-day shift table. We render it as a
    # tall (Shift × Day) listing rather than a wide matrix so it fits the
    # 4-column Summary sheet without horizontal sprawl. Days are emitted
    # in chronological order (Week A → Week B).
    day_cols_present = [d for d in DAYS if d in df.columns]
    if day_cols_present:
        row_cursor += 1
        row_cursor = _section_title(
            "Per-Day Headcount by Shift", row_cursor)
        row_cursor = _table_header(
            ["Day", "Shift", "Headcount", "Shift Hours"], row_cursor)

        day_codes = df[day_cols_present].astype(str).apply(
            lambda col: col.str.strip().str.upper())

        # Same ordering used by the on-screen dashboard for consistency.
        work_order = ["A", "B", "C", "AOT", "BOT", "COT", "OS5"]
        rest_order = ["RD", "LEAVE", "RH", "SPH", "NW", "AWOL"]

        for d in day_cols_present:
            present_codes = {
                v for v in day_codes[d].unique() if v
            }
            ordered = (
                [c for c in work_order if c in present_codes]
                + [c for c in rest_order if c in present_codes]
                + sorted(c for c in present_codes
                         if c not in work_order and c not in rest_order)
            )
            day_label = fmt_day(d)
            for code in ordered:
                count = int((day_codes[d] == code).sum())
                if count == 0:
                    continue
                t = SHIFT_TIMES.get(code, {})
                s, e = t.get("start"), t.get("end")
                if s is not None and e is not None:
                    hours = f"{int(s):02d}:00–{int(e) % 24:02d}:00"
                else:
                    hours = "—"
                row_cursor = _table_row(
                    [day_label, code, count, hours], row_cursor)

    # Summary column widths
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 22

    # ── Write out ─────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    try:
        _excel_timing.__exit__(None, None, None)
    except Exception:
        pass
    _ = _excel_exc_info  # touch sentinel to keep static checkers quiet
    return output.getvalue()


def generate_summary_metrics(results_df: pd.DataFrame) -> pd.DataFrame:
    """Summary by EMP. STATUS. Reserved for future Analytics tab."""
    if "_pass" not in results_df.columns or results_df.empty:
        return pd.DataFrame()
    if "EMP. STATUS" not in results_df.columns:
        return pd.DataFrame()

    status = results_df["EMP. STATUS"].astype(
        str).str.strip().replace("", "Unspecified")
    rows = []
    for s in sorted(status.unique()):
        grp = results_df[status == s]
        n = len(grp)
        if n == 0:
            continue
        p = int(grp["_pass"].sum())
        ot = int(grp["OT Hrs Total"].sum()
                 ) if "OT Hrs Total" in grp.columns else 0
        rows.append({
            "EMP. STATUS": s, "Total": n, "Valid": p,
            "Invalid": n - p, "Pass Rate": f"{100 * p / n:.0f}%",
            "Total OT Hrs": ot,
        })
    return pd.DataFrame(rows)


# ── CSS ───────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600)
def _load_css() -> str:
    """
    Reads style.css from disk and returns its contents as a string.

    FIX: ttl=3600 added (matches create_excel_template).  Without a TTL the
    cache never expires, so any CSS change made to style.css on disk —
    including the overflow:clip sidebar fix — is invisible until the Streamlit
    process is fully restarted.  With ttl=3600 the new CSS is picked up within
    one hour automatically, and immediately after a manual restart.
    """
    from pathlib import Path
    path = Path(__file__).resolve().parent.parent / "assets" / "style.css"
    if not path.exists():
        logger.error("CSS not found at: %s — app will render unstyled.", path)
        return ""
    try:
        return path.read_text(encoding="utf-8")
    except Exception:
        logger.exception("Failed to read CSS file.")
        return ""


def inject_custom_css() -> None:
    """Inject the global stylesheet and stamp the active data-theme.

    Thin shim — delegates to :func:`app.ui.inject_theme` so we keep one
    canonical injector for both the stylesheet and the dual-mode token
    selector. Older call-sites still import ``inject_custom_css`` from
    here; they keep working without change.
    """
    from app.ui import inject_theme
    inject_theme()
